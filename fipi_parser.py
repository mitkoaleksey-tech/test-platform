"""
fipi_parser.py — Модуль парсинга, очистки и валидации данных из официального банка ФИПИ.

Этот модуль полностью независим от основной платформы Reshaemo и работает
исключительно с датасетом ФИПИ (источник 'fipi'). SdamGIA игнорируется.

Входные данные:
    - archive/tasks.csv          — таблица всех заданий
    - archive/free_response.csv  — задания с развернутым ответом
    - archive/rubrics.csv        — критерии и уровни оценивания
    - archive/assets/            — директория с изображениями

Зависимости: pandas, pathlib, json (стандартная библиотека)
"""

from __future__ import annotations

import json
import math
import re
import shutil
import sys
import warnings
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Константы
# ---------------------------------------------------------------------------

FIPI_SOURCE: str = "fipi"

# Маппинг русских названий предметов из датасета на enum-коды платформы
SUBJECT_NAME_MAP: dict[str, str] = {
    "Математика. Профильный уровень": "MATHEMATICS_PROF",
    "Математика. Базовый уровень":    "MATHEMATICS_BASE",
    "Математика":                      "MATHEMATICS",
    "Физика":                          "PHYSICS",
    "Химия":                           "CHEMISTRY",
    "Биология":                        "BIOLOGY",
    "Информатика":                     "INFORMATICS",
    "Обществознание":                  "SOCIAL_STUDIES",
    "История":                         "HISTORY",
    "География":                       "GEOGRAPHY",
    "Русский язык":                    "RUSSIAN",
    "Литература":                      "LITERATURE",
    "Английский язык":                 "ENGLISH",
    "Немецкий язык":                   "GERMAN",
    "Французский язык":                "FRENCH",
    "Испанский язык":                  "SPANISH",
    "Китайский язык":                  "CHINESE",
}
# Обратный маппинг: enum-код -> русское название
SUBJECT_CODE_MAP: dict[str, str] = {v: k for k, v in SUBJECT_NAME_MAP.items()}

# Типы данных колонок при загрузке CSV
TASKS_DTYPES: dict[str, str] = {
    "task_number": "string",
    "fipi_id":     "string",
}
FREE_RESPONSE_DTYPES: dict[str, str] = {
    "task_number": "string",
}

# Паттерн LaTeX-формул: одиночные $ ... $ или двойные $$ ... $$
LATEX_PATTERN: re.Pattern = re.compile(r"\${1,2}[^$]+\${1,2}")

# Поддерживаемые расширения изображений
IMAGE_EXTENSIONS: frozenset[str] = frozenset({".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp"})

# Русские названия трёх математических предметов ФИПИ для экспорта разметки
MATH_SUBJECTS_RU: tuple[str, ...] = (
    "Математика",                    # ОГЭ
    "Математика. Базовый уровень",   # ЕГЭ База
    "Математика. Профильный уровень",# ЕГЭ Профиль
)

# Определение типа экзамена по названию предмета
EXAM_TYPE_MAP: dict[str, str] = {
    "Математика":                     "OGE",
    "Математика. Базовый уровень":    "EGE",
    "Математика. Профильный уровень": "EGE",
}

MATH_EXCEL_COLUMNS: tuple[str, ...] = (
    "external_id",
    "subject",
    "exam_type",
    "task_bank",
    "task_number",
    "subtopic",
    "question",
    "image_files",
    "correct_answer",
    "has_detailed_answer",
    "answer_type",
)


# ===========================================================================
# 0. ВСПОМОГАТЕЛЬНЫЕ УТИЛИТЫ
# ===========================================================================

def parse_image_names(cell_value: object, sep: str = ",") -> list[str]:
    """
    Парсит ячейку с именами/URL файлов изображений в список строк.

    Поддерживает все форматы, встречающиеся в датасете ФИПИ:
      - JSON-массив:   '["file1.png", "file2.png"]'
      - Одиночная строка с кавычками: '"file.png"'
      - Простой список: 'file1.png,file2.png'
      - Пустые значения: None, '', '[]', 'nan'

    Args:
        cell_value: Значение ячейки из DataFrame.
        sep:        Разделитель для не-JSON значений (по умолчанию ',').

    Returns:
        Список очищенных имён файлов без лишних кавычек и скобок.
    """
    if cell_value is None:
        return []
    raw = str(cell_value).strip()
    if not raw or raw in ("nan", "[]", "None"):
        return []

    # Попытка 1: парсим как JSON-массив ["file1.png", "file2.png"]
    if raw.startswith("["):
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except (json.JSONDecodeError, ValueError):
            pass
        # JSON сломан — убираем скобки и парсим как CSV
        raw = raw.strip("[]")

    # Попытка 2: разбиваем по разделителю, убираем кавычки из каждой части
    parts = raw.split(sep)
    result: list[str] = []
    for part in parts:
        name = part.strip().strip("\"'").strip()
        if name:
            result.append(name)
    return result


def parse_options(options_raw: object) -> list[str]:
    """
    Парсит ячейку с вариантами ответов (JSON-массив или строка) в список строк.
    """
    if options_raw is None:
        return []
    raw_str = str(options_raw).strip()
    if not raw_str or raw_str in ("[]", "nan", "None"):
        return []
    if raw_str.startswith("["):
        try:
            parsed = json.loads(raw_str)
            if isinstance(parsed, list):
                return [str(x).strip() for x in parsed if str(x).strip()]
        except Exception:
            pass
    return [x.strip().strip("\"'") for x in raw_str.split(",") if x.strip()]


def clean_fipi_question_text(text: str, options_raw: object = None) -> str:
    """
    Очищает и форматирует текст условия задачи из датасета ФИПИ.

    Выполняет:
      1. Очистку сырого HTML, MathJax (<mjx-container>) и KaTeX (<span class="katex">) в формулы LaTeX $...$.
      2. Удаление заглушек-маркеров картинок: [рис.], [рис. 1], (см. рис. 1), [рисунок 1]
      3. Разделение подпунктов вопросов а), б), в), г) переносом на новые строки.
      4. Прикрепление вариантов ответов из колонки 'options' (1), 2), 3)...).
      5. Преобразование многострочных, однострочных и табулированных (\\t) таблиц в HTML <table>.
      6. Очистку остаточных пайпов | и нормализацию переносов строк.

    Args:
        text:        Исходный текст условия задачи.
        options_raw: Варианты ответов из колонки options (JSON или список).

    Returns:
        Очищенный и красиво форматированный текст.
    """
    if not text or not isinstance(text, str):
        text = ""

    cleaned = text

    # 0. Распаковка псевдо-математических HTML таблиц ФИПИ (<table class="fipi-table">)
    def _unwrap_fipi_math_table(m: re.Match) -> str:
        table_html = m.group(0)
        cells_text = re.sub(r'</?(?:table|tr|th|td)[^>]*>', ' ', table_html)
        cells_text = re.sub(r'\s+', ' ', cells_text).strip()
        # Очищаем битый синтаксис ${...}$ и ${\begin...
        cells_text = re.sub(r'^\$\{\s*', '$', cells_text)
        cells_text = re.sub(r'\s*\}\$$', '$', cells_text)
        if not cells_text.startswith('$'):
            cells_text = f"${cells_text}"
        if not cells_text.endswith('$'):
            cells_text = f"{cells_text}$"
        return f" {cells_text} "

    cleaned = re.sub(r'<table[^>]*>\s*<tr>\s*(?:<th[^>]*>.*?</th>|<td[^>]*>.*?</td>)+\s*</tr>\s*</table>', _unwrap_fipi_math_table, cleaned, flags=re.DOTALL | re.IGNORECASE)
    # Нормализация остаточных опечаток ${...}$ в тексте
    cleaned = re.sub(r'\$\{\s*', '$', cleaned)
    cleaned = re.sub(r'\s*\}\$', '$', cleaned)

    # 1. Замена KaTeX <span class="katex">...<annotation encoding="application/x-tex">...
    def _replace_katex(m: re.Match) -> str:
        tex = m.group(1).strip()
        tex = re.sub(r"_{}\^{}", "", tex)
        tex = re.sub(r"_{}^{}", "", tex)
        tex = re.sub(r"\\_{}\^{}", "", tex)
        tex = re.sub(r"\\_{}^{}", "", tex)
        return f" ${tex}$ "

    cleaned = re.sub(r'<span class="katex">.*?<annotation encoding="application/x-tex">(.*?)</annotation>.*?</span>', _replace_katex, cleaned, flags=re.DOTALL)

    # 2. Замена MathJax <mjx-container>...
    def _replace_mjx(m: re.Match) -> str:
        content = m.group(1)
        chars = re.findall(r'<mjx-c[^>]*>(.*?)</mjx-c>', content)
        if chars:
            math_str = "".join(chars)
            math_str = math_str.replace("𝑃", "P").replace("𝐼", "I").replace("𝑅", "R").replace("𝑥", "x").replace("𝑦", "y").replace("𝑎", "a").replace("𝑏", "b")
            return f" ${math_str}$ "
        return ""

    cleaned = re.sub(r'<mjx-container[^>]*>(.*?)</mjx-container>', _replace_mjx, cleaned, flags=re.DOTALL)

    # 3. Удаление скриптов, форм, hidden-инпутов и тегов обёрток
    cleaned = re.sub(r'<script[^>]*>.*?</script>', '', cleaned, flags=re.DOTALL | re.IGNORECASE)
    cleaned = re.sub(r'<form[^>]*>.*?</form>', '', cleaned, flags=re.DOTALL | re.IGNORECASE)

    # 4. Удаление заглушек рисунков: [рис.], [рис. 1], (см. рис.), [рисунок 1]
    cleaned = re.sub(r"\[\s*рис\.?\s*\d*\s*\]", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\(\s*см\.?\s*рис\.?\s*\d*\s*\)", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\[\s*рисунок\s*\d*\s*\]", "", cleaned, flags=re.IGNORECASE)

    # 5. Форматирование подпунктов вопросов: а), б), в), г) переносом на новую строку
    cleaned = re.sub(r"(^|[\.\!\?\;\s])([а-гА-ГА-ВА-В]\))\s*", r"\1\n\2 ", cleaned)

    # 6. Прикрепление вариантов ответов из колонки 'options'
    opts = parse_options(options_raw)
    if opts:
        first_opt_sample = opts[0][:10] if len(opts[0]) > 5 else opts[0]
        if first_opt_sample not in cleaned:
            opt_lines = []
            for idx, opt in enumerate(opts, start=1):
                if re.match(r"^\d+[\.\)]", opt):
                    opt_lines.append(opt)
                else:
                    opt_lines.append(f"{idx}) {opt}")
            cleaned += "\n\n" + "\n".join(opt_lines)

    # 7. Преобразование таблиц с пайпами | и табуляциями \t в HTML <table>
    cleaned = convert_pipes_to_html_table(cleaned)

    # 8. Удаление остаточных одиночных символов '|' вне таблиц
    cleaned = re.sub(r"^\s*\|\s*$", "", cleaned, flags=re.MULTILINE)
    cleaned = re.sub(r"(\n\s*\|\s*\n)", "\n", cleaned)

    # 9. Очистка остаточного HTML-мусора
    cleaned = re.sub(r'</?(?:div|span|p|a|input|script)[^>]*>', ' ', cleaned, flags=re.IGNORECASE)

    # 10. Нормализация двойных переносов строк и пробелов
    cleaned = re.sub(r"[ \t]+", " ", cleaned)
    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)

    # 11. Очистка пустых хвостовых маркеров ответов 1)\n2)\n3)\n4), когда варианты - это картинки
    cleaned = re.sub(r"(?:\n\s*[1-4][\.\)]\s*){2,4}\s*$", "", cleaned)
    return cleaned.strip()


def convert_pipes_to_html_table(text: str) -> str:
    """
    Преобразует многострочные (2, 3, N колонок), однострочные и табулированные (\\t) таблицы в HTML-таблицы.
    """
    if not text or ("|" not in text and "\t" not in text):
        return text

    lines = [line.strip() for line in text.splitlines()]
    result_lines = []
    i = 0
    in_table = False
    is_first_row = True

    while i < len(lines):
        line = lines[i]

        if not line:
            if in_table:
                has_next_pipe = False
                for j in range(i + 1, min(i + 5, len(lines))):
                    if lines[j] == "|" or "|" in lines[j] or "\t" in lines[j] or lines[j].endswith(":") or "пакеты" in lines[j].lower():
                        has_next_pipe = True
                        break
                if has_next_pipe:
                    i += 1
                    continue
                else:
                    result_lines.append("</table>\n")
                    in_table = False
                    is_first_row = True
            i += 1
            continue

        # Таблица на табуляциях (\t)
        if "\t" in line:
            cells = [c.strip() for c in line.split("\t") if c.strip()]
            if len(cells) >= 2:
                if not in_table:
                    in_table = True
                    is_first_row = True
                    result_lines.append('<table class="fipi-table">')

                tag = "th" if is_first_row else "td"
                row_str = "  <tr>" + "".join(f"<{tag}>{c}</{tag}>" for c in cells) + "</tr>"
                result_lines.append(row_str)
                is_first_row = False
                i += 1
                continue

        # Многострочная последовательность с пайпами: Ячейка 1 \n | \n Ячейка 2 (\n | \n Ячейка 3 ...)
        if i + 2 < len(lines) and lines[i + 1] == "|" and line and line != "|":
            row_cells = [line]
            curr = i + 1
            while curr < len(lines) and lines[curr] == "|":
                if curr + 1 < len(lines) and lines[curr + 1] and lines[curr + 1] != "|":
                    row_cells.append(lines[curr + 1])
                    curr += 2
                else:
                    curr += 1

            if len(row_cells) >= 2:
                if not in_table:
                    in_table = True
                    is_first_row = True
                    result_lines.append('<table class="fipi-table">')

                tag = "th" if is_first_row else "td"
                cells_html = "".join(f"<{tag}>{c}</{tag}>" for c in row_cells)
                result_lines.append(f"  <tr>{cells_html}</tr>")
                is_first_row = False
                i = curr
                continue

        # Заголовок внутри открытой таблицы (например "В абонентскую плату включены пакеты:")
        if in_table and (line.endswith(":") or "пакеты" in line.lower() or "расходования" in line.lower()):
            result_lines.append(f'  <tr><td colspan="4" style="font-weight:bold; background:var(--bg-hover, #f1f5f9); text-align:center; padding: 6px;">{line}</td></tr>')
            i += 1
            continue

        # Однострочная таблица с несколькими ячейками "Cell1 | Cell2 | Cell3"
        if "|" in line:
            cells = [c.strip() for c in line.split("|") if c.strip()]
            if len(cells) >= 2:
                if not in_table:
                    in_table = True
                    is_first_row = True
                    result_lines.append('<table class="fipi-table">')

                tag = "th" if is_first_row else "td"
                row_str = "  <tr>" + "".join(f"<{tag}>{c}</{tag}>" for c in cells) + "</tr>"
                result_lines.append(row_str)
                is_first_row = False
                i += 1
                continue
            elif len(cells) == 1:
                if not in_table:
                    in_table = True
                    is_first_row = True
                    result_lines.append('<table class="fipi-table">')
                result_lines.append(f'  <tr><td colspan="4" style="font-weight:bold; text-align:center;">{cells[0]}</td></tr>')
                i += 1
                continue

        # Обычный текст — закрываем открытую таблицу
        if in_table:
            result_lines.append("</table>\n")
            in_table = False
            is_first_row = True

        result_lines.append(line)
        i += 1

    if in_table:
        result_lines.append("</table>")

    return "\n".join(result_lines)


# ===========================================================================
# 1. ЗАГРУЗКА И ТИПИЗАЦИЯ ДАННЫХ
# ===========================================================================

def load_tasks(csv_path: Path) -> pd.DataFrame:
    """
    Загружает таблицу заданий из tasks.csv с правильными типами данных.

    Args:
        csv_path: Путь к файлу tasks.csv.

    Returns:
        DataFrame всех заданий (включая не-ФИПИ; фильтрация — в других функциях).

    Raises:
        FileNotFoundError: Если файл не найден по указанному пути.
        ValueError: Если файл имеет неожиданную структуру.
    """
    if not csv_path.exists():
        raise FileNotFoundError(f"Файл заданий не найден: {csv_path}")

    try:
        df = pd.read_csv(
            csv_path,
            low_memory=False,
            dtype=TASKS_DTYPES,
        )
    except Exception as exc:
        raise ValueError(f"Ошибка загрузки {csv_path}: {exc}") from exc

    # Нормализуем: trim строковые колонки (совместимо с Pandas 2 и 3)
    str_cols = df.select_dtypes(include=["object", "string"]).columns
    for col in str_cols:
        try:
            df[col] = df[col].str.strip()
        except AttributeError:
            pass

    print(f"[load_tasks] Загружено строк: {len(df):,}  |  колонки: {list(df.columns)}", flush=True)
    return df


def load_free_response(csv_path: Path) -> pd.DataFrame:
    """
    Загружает таблицу заданий с развернутым ответом из free_response.csv.

    Args:
        csv_path: Путь к файлу free_response.csv.

    Returns:
        DataFrame заданий с развернутым ответом.

    Raises:
        FileNotFoundError: Если файл не найден по указанному пути.
        ValueError: Если файл имеет неожиданную структуру.
    """
    if not csv_path.exists():
        raise FileNotFoundError(f"Файл развернутых ответов не найден: {csv_path}")

    try:
        df = pd.read_csv(csv_path, dtype=FREE_RESPONSE_DTYPES)
    except Exception as exc:
        raise ValueError(f"Ошибка загрузки {csv_path}: {exc}") from exc

    str_cols = df.select_dtypes(include=["object", "string"]).columns
    for col in str_cols:
        try:
            df[col] = df[col].str.strip()
        except AttributeError:
            pass

    print(f"[load_free_response] Загружено строк: {len(df):,}", flush=True)
    return df


def load_rubrics(csv_path: Path) -> pd.DataFrame:
    """
    Загружает таблицу официальных критериев и уровней оценивания из rubrics.csv.

    Args:
        csv_path: Путь к файлу rubrics.csv.

    Returns:
        DataFrame рубрик (критериев и уровней оценивания).

    Raises:
        FileNotFoundError: Если файл не найден по указанному пути.
    """
    if not csv_path.exists():
        raise FileNotFoundError(f"Файл рубрик не найден: {csv_path}")

    try:
        df = pd.read_csv(csv_path)
    except Exception as exc:
        raise ValueError(f"Ошибка загрузки {csv_path}: {exc}") from exc

    print(f"[load_rubrics] Загружено строк: {len(df):,}", flush=True)
    return df


# ===========================================================================
# 2. ФИЛЬТРАЦИЯ ПО ИСТОЧНИКУ: СТРОГИЙ ФИЛЬТР ФИПИ
# ===========================================================================

def _apply_fipi_strict_filter(df: pd.DataFrame) -> pd.DataFrame:
    """
    Внутренняя функция. Фильтрует DataFrame строго по источнику 'fipi'.
    SdamGIA и любые другие источники полностью игнорируются.

    Args:
        df: Исходный DataFrame с колонкой 'source'.

    Returns:
        Отфильтрованный DataFrame (только source == 'fipi').
    """
    if "source" not in df.columns:
        warnings.warn(
            "[STRICT FIPI] Колонка 'source' отсутствует — фильтрация по источнику невозможна.",
            stacklevel=3,
        )
        return df

    fipi_df = df[df["source"] == FIPI_SOURCE].copy()
    excluded = len(df) - len(fipi_df)
    if excluded > 0:
        print(
            f"[STRICT FIPI] Исключено {excluded:,} не-ФИПИ строк. "
            f"Итого ФИПИ: {len(fipi_df):,}",
            flush=True,
        )
    return fipi_df


# ===========================================================================
# 3. ОБРАБОТКА ФОРМУЛ (LaTeX / MathML → LaTeX)
# ===========================================================================

def validate_latex_in_text(text: str) -> dict[str, object]:
    """
    Валидирует наличие и корректность LaTeX-формул в тексте.
    ФИПИ использует LaTeX, сконвертированный из исходного MathML.

    Args:
        text: Строка текста для проверки.

    Returns:
        Словарь {
            'has_latex': bool,           # наличие хотя бы одной формулы
            'count': int,                # количество найденных формул
            'formulas': list[str],       # список найденных формул
            'unclosed_tags': bool,       # наличие незакрытых $ символов
        }
    """
    if not isinstance(text, str) or not text.strip():
        return {"has_latex": False, "count": 0, "formulas": [], "unclosed_tags": False}

    formulas = LATEX_PATTERN.findall(text)

    # Исключаем валюту ($40,000, $5,000) и адресацию Excel (B$3, $C$2), чтобы избежать ложных срабатываний
    clean_text = re.sub(r"\$\d[\d,.]*", "", text)
    clean_text = re.sub(r"[A-Za-z]\$\d+", "", clean_text)
    clean_text = re.sub(r"\$[A-Za-z]\$\d+", "", clean_text)
    clean_text = re.sub(r"\$[A-Za-z]\b", "", clean_text)

    # Проверка незакрытых одиночных $: нечётное количество символов $
    dollar_count = clean_text.count("$")
    # Компенсируем парные $$ (каждая пара $$ = 2 символа $)
    double_dollars = len(re.findall(r"\$\$", clean_text)) * 2
    single_dollars = dollar_count - double_dollars
    unclosed = (single_dollars % 2) != 0

    return {
        "has_latex": bool(formulas),
        "count": len(formulas),
        "formulas": formulas,
        "unclosed_tags": unclosed,
    }


def validate_latex_column(df: pd.DataFrame, column: str = "question") -> pd.DataFrame:
    """
    Добавляет в DataFrame колонки с результатами валидации LaTeX-формул.

    Args:
        df:     Исходный DataFrame.
        column: Название колонки с текстом задания (по умолчанию 'question').

    Returns:
        DataFrame с новыми колонками:
            - latex_has_formula  (bool)
            - latex_formula_count (int)
            - latex_unclosed (bool)
    """
    if column not in df.columns:
        warnings.warn(f"[LaTeX] Колонка '{column}' не найдена в DataFrame.", stacklevel=2)
        return df

    results = df[column].fillna("").apply(validate_latex_in_text)

    df = df.copy()
    df["latex_has_formula"]    = results.apply(lambda r: r["has_latex"])
    df["latex_formula_count"]  = results.apply(lambda r: r["count"])
    df["latex_unclosed"]       = results.apply(lambda r: r["unclosed_tags"])

    broken = df["latex_unclosed"].sum()
    if broken > 0:
        warnings.warn(
            f"[LaTeX] Обнаружено {broken} заданий с незакрытыми $ в колонке '{column}'.",
            stacklevel=2,
        )

    return df


# ===========================================================================
# 4. РАБОТА С МЕДИАФАЙЛАМИ (Изображения)
# ===========================================================================

def validate_image_files(
    df: pd.DataFrame,
    assets_dir: Optional[Path],
    image_col: str = "image_files",
    url_col: str = "images",
    sep: str = ",",
) -> pd.DataFrame:
    """
    Проверяет наличие изображений для заданий.

    Работает в двух режимах:
    - LOCAL: если assets_dir существует — проверяет физические файлы в папке.
    - URL:   если assets_dir отсутствует — подсчитывает URL-ссылки из колонки 'images'.

    Примечание: в оригинальном датасете ФИПИ изображения хранятся как URL (колонка 'images'),
    а не как локальные файлы в image_files.

    Args:
        df:         DataFrame с колонками для изображений.
        assets_dir: Путь к директории assets/ или None для URL-режима.
        image_col:  Колонка с именами локальных файлов (по умолчанию 'image_files').
        url_col:    Колонка с URL-ссылками на изображения (по умолчанию 'images').
        sep:        Разделитель имён/URL в ячейке.

    Returns:
        DataFrame с новыми колонками:
            - images_declared (list[str])   — список объявленных файлов/URL
            - images_missing  (list[str])   — файлы не найденные локально (только LOCAL-режим)
            - images_valid    (bool)        — всегда True в URL-режиме
            - images_mode     (str)         — 'local' или 'url'
    """
    df = df.copy()
    # LOCAL-режим только если папка существует И содержит хотя бы один файл
    local_mode = (
        assets_dir is not None
        and assets_dir.exists()
        and any(True for f in assets_dir.iterdir() if f.suffix.lower() in IMAGE_EXTENSIONS)
    )

    if local_mode:
        # LOCAL-режим: проверяем физические файлы
        existing_files: set[str] = {
            f.name.lower()
            for f in assets_dir.iterdir()  # type: ignore[union-attr]
            if f.suffix.lower() in IMAGE_EXTENSIONS
        }
        col_to_use = image_col if image_col in df.columns else url_col

        def _check_local(cell_value: object) -> dict[str, object]:
            if pd.isna(cell_value) or str(cell_value).strip() == "":
                return {"images_declared": [], "images_missing": [], "images_valid": True}
            declared = [n.strip() for n in str(cell_value).split(sep) if n.strip()]
            missing = [n for n in declared if n.lower() not in existing_files]
            return {"images_declared": declared, "images_missing": missing, "images_valid": len(missing) == 0}

        checked = df[col_to_use].apply(_check_local) if col_to_use in df.columns else pd.Series([{"images_declared": [], "images_missing": [], "images_valid": True}] * len(df))
        mode_label = "local"
    else:
        # URL-режим: изображения хранятся как URL, локальной папки нет
        col_to_use = url_col if url_col in df.columns else image_col

        def _check_url(cell_value: object) -> dict[str, object]:
            if pd.isna(cell_value) or str(cell_value).strip() in ("", "[]", "nan"):
                return {"images_declared": [], "images_missing": [], "images_valid": True}
            # URL или список через запятую
            raw = str(cell_value).strip().strip("[]").strip("'\"")
            urls = [u.strip().strip("'\"'") for u in raw.split(",") if u.strip()]
            return {"images_declared": urls, "images_missing": [], "images_valid": True}

        checked = df[col_to_use].apply(_check_url) if col_to_use in df.columns else pd.Series([{"images_declared": [], "images_missing": [], "images_valid": True}] * len(df))
        mode_label = "url"

    df["images_declared"] = checked.apply(lambda r: r["images_declared"])
    df["images_missing"]  = checked.apply(lambda r: r["images_missing"])
    df["images_valid"]    = checked.apply(lambda r: r["images_valid"])
    df["images_mode"]     = mode_label

    total_with_images = (df["images_declared"].apply(len) > 0).sum()
    total_broken      = (~df["images_valid"]).sum()
    print(
        f"[Images] Режим: {mode_label.upper()}  |  "
        f"Заданий с картинками: {total_with_images:,}  |  "
        f"Отсутствующих локально: {total_broken:,}",
        flush=True,
    )
    return df


# ===========================================================================
# 5. КОДИФИКАТОРЫ ФИПИ: ИЕРАРХИЯ ТЕМ
# ===========================================================================

def build_codifier_hierarchy(df_tasks: pd.DataFrame) -> dict[str, list[str]]:
    """
    Строит иерархию тем по официальным кодификаторам КЭС ФИПИ.
    Использует колонки 'codifier_codes' и 'codifier_topics' (если присутствуют).

    Args:
        df_tasks: DataFrame ФИПИ-заданий (уже отфильтрованный).

    Returns:
        Словарь {codifier_code: [subtopic, ...]} — иерархия тем.
    """
    hierarchy: dict[str, list[str]] = {}

    has_codes  = "codifier_codes"  in df_tasks.columns
    has_topics = "codifier_topics" in df_tasks.columns

    if not has_codes:
        warnings.warn("[Codifier] Колонка 'codifier_codes' не найдена.", stacklevel=2)
        return hierarchy

    for _, row in df_tasks.iterrows():
        raw_codes = row.get("codifier_codes", None)
        if pd.isna(raw_codes):
            continue

        codes = [c.strip() for c in str(raw_codes).split(",") if c.strip()]
        topic = str(row.get("codifier_topics", "")).strip() if has_topics else ""

        for code in codes:
            if code not in hierarchy:
                hierarchy[code] = []
            if topic and topic not in hierarchy[code]:
                hierarchy[code].append(topic)

    print(f"[Codifier] Уникальных кодов КЭС: {len(hierarchy)}", flush=True)
    return hierarchy


# ===========================================================================
# 6. ГОТОВЫЕ ПАЙПЛАЙНЫ ФИПИ
# ===========================================================================

def get_fipi_tasks(df_tasks: pd.DataFrame) -> pd.DataFrame:
    """
    Возвращает чистую выборку официальных заданий ФИПИ без дубликатов.

    Применяет:
        - Строгий фильтр по source == 'fipi'
        - Исключение дубликатов: duplicate_of.isna()
        - Валидацию LaTeX-формул

    Args:
        df_tasks: Исходный DataFrame всех заданий (из tasks.csv).

    Returns:
        Чистый DataFrame только ФИПИ-заданий без дублей.
    """
    df = _apply_fipi_strict_filter(df_tasks)

    if "duplicate_of" in df.columns:
        before = len(df)
        df = df[df["duplicate_of"].isna()].copy()
        removed = before - len(df)
        if removed > 0:
            print(f"[get_fipi_tasks] Удалено дублей: {removed:,}", flush=True)
    else:
        warnings.warn("[get_fipi_tasks] Колонка 'duplicate_of' не найдена.", stacklevel=2)

    df = df.reset_index(drop=True)

    # Валидация LaTeX в тексте задания
    question_col = "question" if "question" in df.columns else None
    if question_col:
        df = validate_latex_column(df, column=question_col)

    # Добавление флага задачи с развернутым ответом
    def _check_detailed(row: pd.Series) -> bool:
        ans_type = str(row.get("answer_type", "")).lower() if "answer_type" in row else ""
        if ans_type in ("long", "extended", "free_response", "detailed"):
            return True
        has_det = row.get("has_detailed_answer", False) if "has_detailed_answer" in row else False
        return str(has_det).lower() in ("true", "1", "yes")

    df["has_detailed_answer"] = df.apply(_check_detailed, axis=1)

    print(f"[get_fipi_tasks] Итоговых чистых ФИПИ-заданий: {len(df):,}", flush=True)
    return df


def get_fipi_free_response(
    df_fr: pd.DataFrame,
    df_ru: pd.DataFrame,
    join_key: str = "task_id",
) -> pd.DataFrame:
    """
    Объединяет задания с развернутым ответом ФИПИ с их рубриками.

    ВАЖНО: Официальный банк ФИПИ не публикует ответы в открытом доступе
    для большинства заданий с развернутым ответом. Поэтому после объединения
    добавляется маркер 'has_rubric' — признак наличия опубликованной рубрики.

    Args:
        df_fr:    DataFrame развернутых ответов (из free_response.csv).
        df_ru:    DataFrame рубрик (из rubrics.csv).
        join_key: Ключевая колонка для LEFT JOIN (по умолчанию 'task_id').

    Returns:
        DataFrame с объединёнными данными и колонкой 'has_rubric'.
    """
    df_fr_fipi = _apply_fipi_strict_filter(df_fr)

    if join_key not in df_fr_fipi.columns:
        warnings.warn(
            f"[get_fipi_free_response] Колонка '{join_key}' отсутствует в free_response.csv.",
            stacklevel=2,
        )
        df_fr_fipi["has_rubric"] = False
        return df_fr_fipi

    if join_key not in df_ru.columns:
        warnings.warn(
            f"[get_fipi_free_response] Колонка '{join_key}' отсутствует в rubrics.csv.",
            stacklevel=2,
        )
        df_fr_fipi["has_rubric"] = False
        return df_fr_fipi

    # LEFT JOIN: сохраняем все задания ФИПИ, добавляем рубрику там, где она есть
    merged = df_fr_fipi.merge(
        df_ru,
        on=join_key,
        how="left",
        suffixes=("", "_rubric"),
    )

    # Считаем рубрику присутствующей, если есть хотя бы одна ненулевая колонка рубрики
    rubric_cols = [c for c in merged.columns if c.endswith("_rubric") or c in ("rubric", "criteria", "max_score")]
    if rubric_cols:
        merged["has_rubric"] = merged[rubric_cols[0]].notna()
    else:
        merged["has_rubric"] = False

    with_rubric    = merged["has_rubric"].sum()
    without_rubric = len(merged) - with_rubric
    print(
        f"[get_fipi_free_response] Всего развернутых ФИПИ: {len(merged):,}  |  "
        f"С рубрикой: {with_rubric:,}  |  Без рубрики (held-out): {without_rubric:,}",
        flush=True,
    )
    return merged


def filter_fipi_by_subject(
    df_tasks: pd.DataFrame,
    subject_name: str,
    subject_col: str = "subject",
) -> pd.DataFrame:
    """
    Фильтрует чистые ФИПИ-задания по предмету.

    Принимает как enum-код ('MATHEMATICS_PROF'), так и русское название
    ('Математика. Профильный уровень') — маппинг выполняется автоматически.

    Args:
        df_tasks:     DataFrame ФИПИ-заданий.
        subject_name: Предмет: enum-код ('MATHEMATICS_PROF') или русское название.
        subject_col:  Название колонки с предметом (по умолчанию 'subject').

    Returns:
        Отфильтрованный DataFrame по указанному предмету.
    """
    df = _apply_fipi_strict_filter(df_tasks)

    if subject_col not in df.columns:
        warnings.warn(
            f"[filter_fipi_by_subject] Колонка '{subject_col}' не найдена.",
            stacklevel=2,
        )
        return df

    # Определяем, что передали: enum-код или русское название
    # Пробуем перевести enum -> русское (для датасета с русскими названиями)
    ru_name = SUBJECT_CODE_MAP.get(subject_name.upper(), None)
    if ru_name is None:
        # Возможно передали само русское название
        ru_name = subject_name

    # Сначала точное совпадение по русскому названию
    mask = df[subject_col] == ru_name
    # Если не нашли — пробуем регистронезависимое совпадение с enum-кодом
    if mask.sum() == 0:
        mask = df[subject_col].str.upper() == subject_name.upper()

    result = df[mask].reset_index(drop=True)
    print(
        f"[filter_fipi_by_subject] Предмет '{subject_name}' -> '{ru_name}': {len(result):,} заданий",
        flush=True,
    )
    return result


# ===========================================================================
# 7. ВСПОМОГАТЕЛЬНАЯ ФУНКЦИЯ: СВОДНЫЙ ОТЧЁТ
# ===========================================================================

def generate_fipi_report(
    df_tasks: pd.DataFrame,
    assets_dir: Optional[Path] = None,
) -> dict[str, object]:
    """
    Генерирует сводный JSON-отчёт по чистым ФИПИ-заданиям.

    Args:
        df_tasks:   DataFrame чистых ФИПИ-заданий (из get_fipi_tasks).
        assets_dir: Опциональная директория assets/ для проверки изображений.

    Returns:
        Словарь со статистикой датасета.
    """
    report: dict[str, object] = {
        "total_fipi_tasks": len(df_tasks),
        "subjects": {},
        "task_number_distribution": {},
        "latex_stats": {},
        "image_stats": {},
    }

    # Разбивка по предметам
    if "subject" in df_tasks.columns:
        report["subjects"] = df_tasks["subject"].value_counts().to_dict()

    # Разбивка по номерам заданий КИМ
    if "task_number" in df_tasks.columns:
        report["task_number_distribution"] = (
            df_tasks["task_number"].value_counts().sort_index().to_dict()
        )

    # LaTeX-статистика
    if "latex_has_formula" in df_tasks.columns:
        report["latex_stats"] = {
            "tasks_with_latex": int(df_tasks["latex_has_formula"].sum()),
            "tasks_with_unclosed_tags": int(df_tasks["latex_unclosed"].sum()),
        }

    # Статистика изображений
    if assets_dir is not None and "image_files" in df_tasks.columns:
        validated = validate_image_files(df_tasks, assets_dir)
        report["image_stats"] = {
            "tasks_with_images": int((validated["images_declared"].apply(len) > 0).sum()),
            "tasks_with_missing_files": int((~validated["images_valid"]).sum()),
        }

    return report


# ===========================================================================
# 8. ЭКСПОРТ МАТЕМАТИЧЕСКИХ ЗАДАЧ ФИПИ В EXCEL ДЛЯ РУЧНОЙ РАЗМЕТКИ
# ===========================================================================

MATH_EXCEL_COLUMNS: list[str] = [
    "external_id",   # ID источника ФИПИ
    "subject",       # Название предмета
    "exam_type",     # EGE / OGE
    "task_bank",     # Всегда FIPI
    "task_number",   # ПУСТО — ручная разметка
    "subtopic",      # ПУСТО — ручная разметка
    "question",      # Текст задачи (LaTeX)
    "image_files",   # Ссылки на картинки через запятую
    "correct_answer",# Ответ, если есть
    "has_detailed_answer", # Признак развёрнутого ответа
    "answer_type",   # Тип ответа из датасета
]


def get_math_fipi_for_export(df_tasks: pd.DataFrame) -> pd.DataFrame:
    """
    Отфильтровывает из общего датасета только три математических предмета ФИПИ
    без дубликатов, готовые к экспорту.

    Args:
        df_tasks: Исходный DataFrame из load_tasks().

    Returns:
        Отфильтрованный DataFrame: только FIPI, только три математики, без дублей.
    """
    # 1. Строгий фильтр ФИПИ
    df = _apply_fipi_strict_filter(df_tasks)

    # 2. Только три математических предмета
    if "subject" not in df.columns:
        raise ValueError("Колонка 'subject' не найдена в датасете.")

    df = df[df["subject"].isin(MATH_SUBJECTS_RU)].copy()
    print(
        f"[MathExport] После фильтра математики: {len(df):,} задач  "
        f"({', '.join(df['subject'].value_counts().to_dict().keys() if len(df) else [])})",
        flush=True,
    )

    # 3. Убираем дубликаты
    if "duplicate_of" in df.columns:
        before = len(df)
        df = df[df["duplicate_of"].isna()].copy()
        print(f"[MathExport] Удалено дублей: {before - len(df):,}", flush=True)

    return df.reset_index(drop=True)


def solve_fipi_math_task(q_raw: str) -> str:
    """
    Математический Python-солвер для типовых задач ФИПИ I части (проценты, пропорции,
    физические формулы, теория вероятностей, показательные и квадратные уравнения).
    """
    if not isinstance(q_raw, str) or not q_raw.strip():
        return ""

    c = re.sub(r'<[^>]+>', '', q_raw)
    c = re.sub(r'\[\s*рис\.?\s*\d*\s*\]', '', c, flags=re.IGNORECASE)
    c = re.sub(r'\s+', ' ', c).strip()

    # 1.1 Автобусы / Экскурсии (округление вверх)
    m = re.search(r'(\d+)\s*детей\s*и\s*(\d+)\s*воспитател.*?не\s*более\s*(\d+)', c, re.IGNORECASE)
    if m:
        n1, n2, n3 = int(m.group(1)), int(m.group(2)), int(m.group(3))
        return str(math.ceil((n1 + n2) / n3))

    # 1.2 Снижение цены в процентах (телефон, товар)
    m = re.search(r'стоил[а-о]?\s*(\d+(?:[.,]\d+)?)\s*рубл.*?стал[а-о]?\s*стоить\s*(\d+(?:[.,]\d+)?)\s*рубл.*?процент', c, re.IGNORECASE)
    if m:
        n1 = float(m.group(1).replace(',', '.'))
        n2 = float(m.group(2).replace(',', '.'))
        if n1 > 0:
            pct = (n1 - n2) / n1 * 100.0
            return str(int(round(pct))) if abs(pct - round(pct)) < 1e-5 else str(round(pct, 2)).replace('.', ',')

    # 1.3 Повышение цены в процентах
    m = re.search(r'стоил[а-о]?\s*(\d+(?:[.,]\d+)?)\s*рубл.*?повысил.*?на\s*(\d+(?:[.,]\d+)?)\s*процент', c, re.IGNORECASE)
    if m:
        n1 = float(m.group(1).replace(',', '.'))
        pct = float(m.group(2).replace(',', '.'))
        res = n1 * (1.0 + pct / 100.0)
        return str(int(round(res))) if abs(res - round(res)) < 1e-5 else str(round(res, 2)).replace('.', ',')

    # 1.4 Отношение ингредиентов (фарш, смесь)
    m = re.search(r'отношении\s*(\d+)\s*:\s*(\d+).*?процент.*?составляет\s*(\w+)', c, re.IGNORECASE)
    if m:
        n1, n2 = float(m.group(1)), float(m.group(2))
        res = n2 / (n1 + n2) * 100.0
        return str(int(round(res))) if abs(res - round(res)) < 1e-5 else str(round(res, 2)).replace('.', ',')

    # 1.5 Сахар в лагере на человека
    m = re.search(r'(\d+)\s*г\s*сахара.*?(\d+)\s*человек.*?упаковок', c, re.IGNORECASE)
    if m:
        g_per_p = float(m.group(1))
        people = float(m.group(2))
        total_kg = (g_per_p * people) / 1000.0
        return str(math.ceil(total_kg))

    # 2.1 P = I^2 * R  -> R = P / I^2
    m = re.search(r'P\s*=\s*I\^?2\s*R.*?P\s*=\s*(\d+(?:[.,]\d+)?).*?I\s*=\s*(\d+(?:[.,]\d+)?)', c, re.IGNORECASE)
    if not m:
        m = re.search(r'мощность.*?P\s*=\s*(\d+(?:[.,]\d+)?).*?сила тока.*?I\s*=\s*(\d+(?:[.,]\d+)?)', c, re.IGNORECASE)
    if m:
        p_val = float(m.group(1).replace(',', '.'))
        i_val = float(m.group(2).replace(',', '.'))
        if i_val > 0:
            r_val = p_val / (i_val ** 2)
            return str(int(round(r_val))) if abs(r_val - round(r_val)) < 1e-5 else str(round(r_val, 2)).replace('.', ',')

    # 2.2 Закон Гука F = kx -> x = F / k
    m = re.search(r'F\s*=\s*kx.*?F\s*=\s*(\d+(?:[.,]\d+)?).*?k\s*=\s*(\d+(?:[.,]\d+)?)', c, re.IGNORECASE)
    if m:
        f_val = float(m.group(1).replace(',', '.'))
        k_val = float(m.group(2).replace(',', '.'))
        if k_val > 0:
            x_val = f_val / k_val
            return str(int(round(x_val))) if abs(x_val - round(x_val)) < 1e-5 else str(round(x_val, 2)).replace('.', ',')

    # 2.3 Высота перил / трапеция средняя линия l = (h1 + h2) / 2
    m = re.search(r'наименьшая\s*высота.*?(\d+(?:[.,]\d+)?).*?наибольшая\s*высота.*?(\d+(?:[.,]\d+)?)', c, re.IGNORECASE)
    if m:
        h1 = float(m.group(1).replace(',', '.'))
        h2 = float(m.group(2).replace(',', '.'))
        l_val = (h1 + h2) / 2.0
        return str(int(round(l_val))) if abs(l_val - round(l_val)) < 1e-5 else str(round(l_val, 2)).replace('.', ',')

    # 3.1 Жребий (N имен)
    m = re.search(r'((?:[А-Я][а-я]+,\s*)+[А-Я][а-я]+\s*и\s*[А-Я][а-я]+)\s*бросили\s*жребий', c)
    if m:
        names = re.findall(r'[А-Я][а-я]+', m.group(1))
        if names:
            prob = 1.0 / len(names)
            return str(round(prob, 2)).replace('.', ',')

    # 3.2 Билеты (выигрышные / всего)
    m = re.search(r'всего\s*(\d+)\s*билет.*?(\d+)\s*из\s*них\s*выигрышн', c, re.IGNORECASE)
    if m:
        total_b = float(m.group(1))
        win_b = float(m.group(2))
        if total_b > 0:
            prob = win_b / total_b
            return str(round(prob, 2)).replace('.', ',')

    # 3.3 Такси (черные / всего)
    m = re.search(r'всего\s*(\d+)\s*такси.*?(\d+)\s*черн', c, re.IGNORECASE)
    if m:
        total_t = float(m.group(1))
        black_t = float(m.group(2))
        if total_t > 0:
            prob = black_t / total_t
            return str(round(prob, 2)).replace('.', ',')

    # 4.1 Показательное уравнение вида (a/b)^(x-c) = (d1/d2) или a^(x-c) = d
    m = re.search(r'\(\s*\\frac\{(\d+)\}\{(\d+)\}\s*\)\s*\^\{\s*x\s*([-+]\s*\d+)\s*\}\s*=\s*\\frac\{(\d+)\}\{(\d+)\}', c)
    if m:
        a, b = float(m.group(1)), float(m.group(2))
        c_val = float(m.group(3).replace(' ', ''))
        d1, d2 = float(m.group(4)), float(m.group(5))
        base = a / b
        target = d1 / d2
        if base > 0 and target > 0 and base != 1:
            power = math.log(target) / math.log(base)
            x_sol = power - c_val
            return str(int(round(x_sol))) if abs(x_sol - round(x_sol)) < 1e-5 else str(round(x_sol, 2)).replace('.', ',')

    # 4.2 Простые показательные уравнения a^(x-c) = b
    m = re.search(r'(\d+)\s*\^\{\s*x\s*([-+]\s*\d+)\s*\}\s*=\s*(\d+)', c)
    if m:
        base = float(m.group(1))
        c_val = float(m.group(2).replace(' ', ''))
        target = float(m.group(3))
        if base > 0 and target > 0 and base != 1:
            power = math.log(target) / math.log(base)
            x_sol = power - c_val
            return str(int(round(x_sol))) if abs(x_sol - round(x_sol)) < 1e-5 else str(round(x_sol, 2)).replace('.', ',')

    # 4.3 Квадратные уравнения ax^2 + bx + c = 0
    m = re.search(r'(\d+)\s*x\s*\^\s*2\s*([-+]\s*\d+)\s*x\s*([-+]\s*\d+)\s*=\s*0', c)
    if m:
        a = float(m.group(1))
        b = float(m.group(2).replace(' ', ''))
        c_v = float(m.group(3).replace(' ', ''))
        disc = b**2 - 4*a*c_v
        if disc >= 0:
            x1 = (-b + math.sqrt(disc)) / (2*a)
            x2 = (-b - math.sqrt(disc)) / (2*a)
            res = min(x1, x2) if 'меньший' in c.lower() else max(x1, x2)
            return str(int(round(res))) if abs(res - round(res)) < 1e-5 else str(round(res, 2)).replace('.', ',')

    return ""


def export_math_fipi_to_review_excel(
    df_tasks: pd.DataFrame,
    output_path: Path,
) -> None:
    """
    Экспортирует три математических предмета ФИПИ в файл tasks.xlsx
    для последующей ручной разметки в Excel.

    Создаёт таблицу с фиксированными 9 колонками.
    Колонки task_number и subtopic оставлены ПУСТЫМИ для ручного заполнения.

    Args:
        df_tasks:    Исходный DataFrame из load_tasks() (нефильтрованный).
        output_path: Путь для сохранения .xlsx файла.

    Raises:
        ValueError: Если в df_tasks нет необходимых колонок.
        OSError:    Если файл не удаётся записать.
    """
    # --- Шаг 1: Фильтрация ---
    df = get_math_fipi_for_export(df_tasks)

    if df.empty:
        print("[MathExport] ПРЕДУПРЕЖДЕНИЕ: Нет данных для экспорта.", flush=True)
        return

    # --- Шаг 2: Формирование нужных колонок ---
    # external_id: предпочитаем fipi_id, иначе id
    if "fipi_id" in df.columns:
        df["external_id"] = df["fipi_id"].fillna("").astype(str)
    elif "id" in df.columns:
        df["external_id"] = df["id"].fillna("").astype(str)
    else:
        df["external_id"] = ""

    # exam_type: выводим из названия предмета
    df["exam_type"] = df["subject"].map(EXAM_TYPE_MAP).fillna("EGE")

    # task_bank: всегда FIPI
    df["task_bank"] = "FIPI"

    # task_number, subtopic: пустые поля для ручной разметки
    df["task_number"] = df["task_number"].fillna("").astype(str) if "task_number" in df.columns else ""
    df["subtopic"]    = ""
    df["answer_type"] = df["answer_type"].fillna("").astype(str) if "answer_type" in df.columns else ""
    df["has_detailed_answer"] = df.apply(
        lambda r: (
            str(r.get("has_detailed_answer", "")).lower() in ("true", "1")
            or str(r.get("answer_type", "")).lower() in ("long", "extended", "free_response", "detailed")
            or "разверн" in str(r.get("answer_type", "")).lower()
            or "изменить статус" in str(r.get("answer", "")).lower()
            or "разверн" in str(r.get("answer", "")).lower()
            or not str(r.get("answer", "")).strip()
        ),
        axis=1
    )

    # question: очищенный и отформатированный текст задачи (без [рис.], с вариантами ответов и HTML таблицами)
    if "question" in df.columns:
        options_series = df["options"] if "options" in df.columns else pd.Series([None] * len(df))
        df["question"] = [
            clean_fipi_question_text(q, opt)
            for q, opt in zip(df["question"].fillna("").astype(str), options_series)
        ]
    else:
        df["question"] = ""

    # image_files: URL-ссылки через запятую (колонка images из датасета ФИПИ)
    if "image_files" in df.columns:
        df["image_files"] = df["image_files"].fillna("").astype(str)
    elif "images" in df.columns:
        # В датасете ФИПИ картинки хранятся как URL в колонке 'images'
        df["image_files"] = df["images"].fillna("").astype(str)
    else:
        df["image_files"] = ""

    # correct_answer: восстанавливаем ответы из SdamGIA только для задач I части (краткий ответ)
    sdamgia = df_tasks[
        (df_tasks["source"] == "sdamgia") &
        (df_tasks["answer"].notna()) &
        (df_tasks["answer"].astype(str).str.strip() != "") &
        (df_tasks["answer_type"] != "long")
    ]

    fipi_id_map = {}
    for _, r in sdamgia[sdamgia["fipi_id"].notna()].iterrows():
        fipi_id_map[str(r["fipi_id"]).strip().upper()] = str(r["answer"]).strip()

    id_map = {}
    for _, r in sdamgia.iterrows():
        id_map[r["id"]] = str(r["answer"]).strip()

    text_map = {}
    for _, r in sdamgia[sdamgia["question"].notna()].iterrows():
        k = re.sub(r'<[^>]+>', '', str(r["question"]))
        k = re.sub(r'\[\s*рис\.?\s*\d*\s*\]', '', k, flags=re.IGNORECASE)
        k = re.sub(r'\s+', '', k).lower()
        if len(k) > 15 and k not in text_map:
            text_map[k] = str(r["answer"]).strip()

    def _get_fp(t: object) -> str:
        if not isinstance(t, str): return ""
        c = re.sub(r'<[^>]+>', '', t)
        c = re.sub(r'\[\s*рис\.?\s*\d*\s*\]', '', c, flags=re.IGNORECASE)
        tokens = re.findall(r'[а-яА-Я]{3,}|\b\d+(?:[.,]\d+)?\b', c.lower())
        return " ".join(tokens)

    fp_map = {}
    for _, r in sdamgia[sdamgia["question"].notna()].iterrows():
        fp = _get_fp(r["question"])
        if len(fp) > 20 and fp not in fp_map:
            fp_map[fp] = str(r["answer"]).strip()

    def _parse_imgs_list(val: object) -> list[str]:
        if pd.isna(val) or not val: return []
        s = str(val).strip()
        if s.startswith('['):
            try: return [str(x).strip().split('/')[-1] for x in json.loads(s) if str(x).strip()]
            except Exception: pass
        return [x.strip().split('/')[-1] for x in s.split(',') if x.strip()]

    img_list_map = {}
    for _, r in sdamgia.iterrows():
        imgs = _parse_imgs_list(r.get("image_files")) or _parse_imgs_list(r.get("images"))
        if imgs:
            key = ",".join(sorted(imgs))
            if key not in img_list_map:
                img_list_map[key] = str(r["answer"]).strip()

    answers = []
    for _, r in df.iterrows():
        # Задачи второй части (развернутый ответ) оставляем БЕЗ ответа
        ans_type = str(r.get("answer_type", "")).lower()
        if ans_type in ("long", "extended", "free_response"):
            answers.append("")
            continue

        ans = str(r.get("answer", "")).strip() if pd.notna(r.get("answer")) else ""
        if not ans or ans in ("nan", "None"):
            fid = str(r.get("fipi_id", "")).strip().upper()
            dup = r.get("duplicate_of", None)
            q_raw = str(r.get("question", ""))
            q_key = re.sub(r'<[^>]+>', '', q_raw)
            q_key = re.sub(r'\[\s*рис\.?\s*\d*\s*\]', '', q_key, flags=re.IGNORECASE)
            q_key = re.sub(r'\s+', '', q_key).lower()

            imgs = _parse_imgs_list(r.get("image_files")) or _parse_imgs_list(r.get("images"))
            img_key = ",".join(sorted(imgs)) if imgs else ""

            if fid and fid in fipi_id_map:
                ans = fipi_id_map[fid]
            elif pd.notna(dup) and dup in id_map:
                ans = id_map[dup]
            elif q_key and q_key in text_map:
                ans = text_map[q_key]
            elif img_key and img_key in img_list_map:
                ans = img_list_map[img_key]
            elif _get_fp(q_raw) in fp_map:
                ans = fp_map[_get_fp(q_raw)]

            # Если сопоставление со SdamGIA не дало результата, запускаем математический Python-солвер
            if not ans:
                ans = solve_fipi_math_task(q_raw)

        answers.append(ans)

    df["correct_answer"] = answers
    recovered_count = sum(1 for a in answers if a)
    print(f"[MathExport] Восстановлено/вычислено правильных ответов для I части ФИПИ: {recovered_count:,} из {len(df):,}", flush=True)

    # Итоговый срез строго из 9 колонок
    export_df = df[MATH_EXCEL_COLUMNS].copy()

    # --- Шаг 3: Формирование Excel-файла через openpyxl ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Математика ФИПИ"

    # Стили шапки
    HEADER_FILL   = PatternFill(start_color="2D5A8E", end_color="2D5A8E", fill_type="solid")
    MANUAL_FILL   = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    HEADER_FONT   = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    MANUAL_FONT   = Font(name="Calibri", bold=True, color="856404",  size=11)
    THIN_BORDER   = Border(
        left=Side(style="thin"),  right=Side(style="thin"),
        top=Side(style="thin"),   bottom=Side(style="thin"),
    )

    # Заголовки и их описания
    COLUMN_META: dict[str, dict] = {
        "external_id":    {"width": 20,  "manual": False},
        "subject":        {"width": 30,  "manual": False},
        "exam_type":      {"width": 10,  "manual": False},
        "task_bank":      {"width": 12,  "manual": False},
        "task_number":    {"width": 14,  "manual": True},   # ← ручная разметка
        "subtopic":       {"width": 35,  "manual": True},   # ← ручная разметка
        "question":       {"width": 80,  "manual": False},
        "image_files":    {"width": 40,  "manual": False},
        "correct_answer": {"width": 18,  "manual": False},
        "has_detailed_answer": {"width": 22, "manual": False},
        "answer_type":    {"width": 18,  "manual": False},
    }

    # Строка 1: заголовки
    for col_idx, col_name in enumerate(MATH_EXCEL_COLUMNS, start=1):
        meta    = COLUMN_META[col_name]
        cell    = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font   = MANUAL_FONT   if meta["manual"] else HEADER_FONT
        cell.fill   = MANUAL_FILL   if meta["manual"] else HEADER_FILL
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        ws.column_dimensions[get_column_letter(col_idx)].width = meta["width"]

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"  # Зафиксировать шапку при прокрутке

    # Строки данных
    DATA_FONT = Font(name="Calibri", size=10)
    WRAP_ALIGN = Alignment(vertical="top", wrap_text=True)
    NOWRAP_ALIGN = Alignment(vertical="top", wrap_text=False)

    for row_idx, row_data in enumerate(export_df.itertuples(index=False), start=2):
        for col_idx, col_name in enumerate(MATH_EXCEL_COLUMNS, start=1):
            value = getattr(row_data, col_name)
            cell  = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.font   = DATA_FONT
            cell.border = THIN_BORDER
            # Текст задачи переносится по словам, остальные — нет
            cell.alignment = WRAP_ALIGN if col_name == "question" else NOWRAP_ALIGN

        # Высота строки: авто для коротких, фиксированная для длинных вопросов
        question_text = str(getattr(row_data, "question", ""))
        estimated_lines = min(len(question_text) // 80 + 1, 15)
        ws.row_dimensions[row_idx].height = max(15, estimated_lines * 15)

    # --- Шаг 4: Сохранение ---
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        output_path = output_path.parent / (output_path.stem + "_new.xlsx")
        print(f"[MathExport] Исходный файл заблокирован (открыт в Excel). Сохраняем в: {output_path}", flush=True)
        wb.save(output_path)

    total_rows = len(export_df)
    size_kb = output_path.stat().st_size // 1024
    print(
        f"[MathExport] Сохранено: {output_path}  "
        f"({total_rows:,} задач, {size_kb:,} КБ)",
        flush=True,
    )
    print(
        f"[MathExport] Жёлтые колонки 'task_number' и 'subtopic' ожидают ручную разметку.",
        flush=True,
    )


def export_task_images(
    df_tasks: pd.DataFrame,
    assets_dir: Path,
    output_dir: Path,
    image_col: str = "image_files",
    sep: str = ",",
    overwrite: bool = False,
) -> dict[str, int]:
    """
    Сканирует отфильтрованный датасет, находит все файлы изображений
    из колонки image_files и копирует их из assets_dir в output_dir.

    Работает в двух режимах:
      - LOCAL: если assets_dir существует и не пустая — копирует физические файлы.
      - URL:   если assets_dir пуста или отсутствует — отображает URL-ссылки
        и предупреждает, что скачивание не выполняется.

    Args:
        df_tasks:   DataFrame с колонкой изображений (отфильтрованный или полный).
        assets_dir: Путь к исходной директории с файлами (archive/assets/).
        output_dir: Целевая директория для скопированных файлов (например ./images/).
        image_col:  Колонка с именами файлов (по умолчанию 'image_files').
        sep:        Разделитель имён файлов в ячейке (по умолчанию ',').
        overwrite:  Если True — перезаписывает уже существующие файлы.

    Returns:
        Словарь со статистикой: {
            'total_declared': объявлено,
            'copied':         успешно скопировано,
            'skipped':        пропущено (уже существуют, overwrite=False),
            'missing':        не найдено в assets_dir,
            'errors':         ошибки при копировании,
        }
    """
    # --- Проверка режима ---
    local_mode = (
        assets_dir.exists()
        and any(True for f in assets_dir.iterdir() if f.suffix.lower() in IMAGE_EXTENSIONS)
    )

    stats: dict[str, int] = {
        "total_declared": 0,
        "copied":         0,
        "skipped":        0,
        "missing":        0,
        "errors":         0,
    }

    if not local_mode:
        # URL-режим: выводим ссылки, копирование невозможно
        url_col = "images"
        col_to_scan = image_col if image_col in df_tasks.columns else (
            url_col if url_col in df_tasks.columns else None
        )
        if col_to_scan:
            all_urls: list[str] = []
            for cell in df_tasks[col_to_scan].dropna():
                all_urls.extend(parse_image_names(cell, sep=sep))
            stats["total_declared"] = len(set(all_urls))

        print(
            f"[Images] Режим URL: файлы хранятся на сервере ФИПИ, локальное копирование невозможно.",
            flush=True,
        )
        print(
            f"[Images] Уникальных URL в датасете: {stats['total_declared']:,}",
            flush=True,
        )
        return stats

    # --- LOCAL-режим ---
    if image_col not in df_tasks.columns:
        warnings.warn(
            f"[export_task_images] Колонка '{image_col}' не найдена.",
            stacklevel=2,
        )
        return stats

    # Строим индекс файлов в assets_dir (без регистра → быстрый поиск)
    asset_index: dict[str, Path] = {
        f.name.lower(): f
        for f in assets_dir.rglob("*")
        if f.is_file() and f.suffix.lower() in IMAGE_EXTENSIONS
    }
    print(
        f"[Images] assets/ содержит {len(asset_index):,} файлов. "
        f"Целевая папка: {output_dir}",
        flush=True,
    )

    # Собираем уникальный список файлов из всех строк через parse_image_names (дедупликация)
    # parse_image_names корректно обрабатывает JSON-массивы ["file.png"] из датасета ФИПИ
    declared_files: set[str] = set()
    missing_names:  list[str] = []

    for cell in df_tasks[image_col].dropna():
        names = parse_image_names(cell, sep=sep)
        declared_files.update(names)

    stats["total_declared"] = len(declared_files)

    # Создаём целевую папку
    output_dir.mkdir(parents=True, exist_ok=True)

    # Копируем файлы
    for filename in sorted(declared_files):
        src_path = asset_index.get(filename.lower())
        dst_path = output_dir / filename

        if src_path is None:
            # Файл не найден в assets/
            missing_names.append(filename)
            stats["missing"] += 1
            continue

        if dst_path.exists() and not overwrite:
            # Уже существует, пропускаем
            stats["skipped"] += 1
            continue

        try:
            shutil.copy2(src_path, dst_path)
            stats["copied"] += 1
        except OSError as exc:
            warnings.warn(
                f"[export_task_images] Ошибка при копировании '{filename}': {exc}",
                stacklevel=2,
            )
            stats["errors"] += 1

    # --- Логирование результатов ---
    print(
        f"[Images] Итого: Объявлено {stats['total_declared']:,}  |  "
        f"Скопировано {stats['copied']:,}  |  "
        f"Пропущено {stats['skipped']:,}  |  "
        f"Отсутствует {stats['missing']:,}  |  "
        f"Ошибок {stats['errors']:,}",
        flush=True,
    )

    if missing_names:
        # Первые 20 отсутствующих — в предупреждение
        sample = missing_names[:20]
        warnings.warn(
            f"[export_task_images] {stats['missing']} файлов не найдено в assets/:\n"
            + "\n".join(f"  - {n}" for n in sample)
            + (f"\n  ... и ещё {stats['missing'] - 20}" if stats['missing'] > 20 else ""),
            stacklevel=2,
        )

    return stats

if __name__ == "__main__":
    # Ожидаемая структура директорий (рядом с данным скриптом)
    ARCHIVE_DIR = Path("archive")
    ASSETS_DIR  = ARCHIVE_DIR / "assets"

    TASKS_CSV   = ARCHIVE_DIR / "tasks.csv"
    FR_CSV      = ARCHIVE_DIR / "free_response.csv"
    RUBRICS_CSV = ARCHIVE_DIR / "rubrics.csv"

    print("=" * 60)
    print("  FIPI Parser — Reshaemo Dataset Tools")
    print("=" * 60)

    # --- Загрузка данных ---
    try:
        df_all_tasks = load_tasks(TASKS_CSV)
    except FileNotFoundError as e:
        print(f"[ОШИБКА] {e}", file=sys.stderr)
        sys.exit(1)

    try:
        df_fr = load_free_response(FR_CSV)
    except FileNotFoundError:
        print("[ПРЕДУПРЕЖДЕНИЕ] free_response.csv не найден, пропускаем.", flush=True)
        df_fr = pd.DataFrame()

    try:
        df_ru = load_rubrics(RUBRICS_CSV)
    except FileNotFoundError:
        print("[ПРЕДУПРЕЖДЕНИЕ] rubrics.csv не найден, пропускаем.", flush=True)
        df_ru = pd.DataFrame()

    # --- Основной пайплайн ФИПИ ---
    df_fipi = get_fipi_tasks(df_all_tasks)

    # Проверка изображений
    df_fipi = validate_image_files(df_fipi, ASSETS_DIR)

    # Построение иерархии тем
    codifier = build_codifier_hierarchy(df_fipi)

    # Пайплайн для предмета — теперь работает и с enum-кодами, и с русскими названиями
    df_math_prof = filter_fipi_by_subject(df_fipi, "MATHEMATICS_PROF")
    df_math_base = filter_fipi_by_subject(df_fipi, "MATHEMATICS_BASE")
    df_physics   = filter_fipi_by_subject(df_fipi, "PHYSICS")

    # Развернутые ответы с рубриками
    if not df_fr.empty and not df_ru.empty:
        # Определяем ключ для JOIN: ищем общую колонку между free_response и rubrics
        fr_cols = set(df_fr.columns)
        ru_cols = set(df_ru.columns)
        common_cols = fr_cols & ru_cols
        # Приоритет ключей: rubric_id есть и в free_response, и в rubrics
        preferred_keys = ["rubric_id", "id", "task_id", "fipi_id", "external_id"]
        join_key = next((k for k in preferred_keys if k in common_cols), None)

        if join_key:
            print(f"[JOIN] Найден ключ для rubrics: '{join_key}'", flush=True)
            df_free = get_fipi_free_response(df_fr, df_ru, join_key=join_key)
        else:
            print(f"[JOIN] Общих ключей не найдено.", flush=True)
            print(f"       free_response колонки: {sorted(fr_cols)}", flush=True)
            print(f"       rubrics колонки:       {sorted(ru_cols)}", flush=True)
            df_free = pd.DataFrame()
    else:
        df_free = pd.DataFrame()

    # --- Сводный отчёт ---
    report = generate_fipi_report(df_fipi, ASSETS_DIR)
    print("\n" + "=" * 60)
    print("  СВОДНЫЙ ОТЧЁТ ФИПИ")
    print("=" * 60)
    print(json.dumps(report, ensure_ascii=False, indent=2, default=str))

    # --- Экспорт математических задач в Excel для ручной разметки ---
    OUTPUT_XLSX = Path("output") / "tasks.xlsx"
    print("\n" + "=" * 60)
    print("  ЭКСПОРТ МАТЕМАТИКИ В EXCEL")
    print("=" * 60)
    export_math_fipi_to_review_excel(df_all_tasks, OUTPUT_XLSX)

    # --- Копирование картинок для математических задач ---
    OUTPUT_IMAGES = Path("output") / "images"
    print("\n" + "=" * 60)
    print("  КОПИРОВАНИЕ ИЗОБРАЖЕНИЙ")
    print("=" * 60)
    # Передаём уже отфильтрованный DataFrame только с математическими задачами
    df_math = get_math_fipi_for_export(df_all_tasks)
    img_stats = export_task_images(
        df_tasks=df_math,
        assets_dir=ASSETS_DIR,
        output_dir=OUTPUT_IMAGES,
        overwrite=False,
    )

    print("\n" + "=" * 60)
    print("  ГОТОВО")
    print("=" * 60)
    print(f"  Excel-файл для разметки:     {OUTPUT_XLSX.resolve()}")
    print(f"  Изображения:                 {OUTPUT_IMAGES.resolve()}")
    print(f"    - Объявлено:   {img_stats['total_declared']:,}")
    print(f"    - Скопировано:  {img_stats['copied']:,}")
    print(f"    - Пропущено:   {img_stats['skipped']:,} (уже существует)")
    print(f"    - Отсутствует: {img_stats['missing']:,} (нет в assets/)")
    if img_stats.get("missing", 0) > 0:
        print(f"    ⚠️  См. предупреждения выше — эти файлы URL на сервере ФИПИ.")
    print(f"  Откройте файл в Excel и заполните жёлтые колонки:")
    print(f"    - task_number  (номер задания в КИМ, например: 1, 2, 3...)")
    print(f"    - subtopic     (название подтемы, например: 'Уравнения')")
    print("=" * 60)

